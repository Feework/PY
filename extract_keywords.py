from collections import Counter
from openpyxl import Workbook
from jieba import analyse

# 打开文件
filename = "data_cut.txt"
f = open(filename, 'r', encoding='utf-8')

#使用analyse提取句子级的关键字，extract_tags提取关键字原理为使用tfidf算法
# 计算tfidf权重并保存
# 此处算法错误，tf-idf(t,d) = tf(t,d) * idf(t,D) ,tf使用的是d为某道题，idf使用的D为某门题库，此处tf用的是题库，idf用的是jieba通用，不准确。
# 应该不同题中相同的词的tf-idf的值是不同的
tfidf = analyse.extract_tags
bin_content = open(filename, 'rb').read()
#此处idf的值应该要计算题库自身的idf模型，这里使用的是通用的不准确；读入的文本应该是某道题而不是整个题库
keywords = tfidf(bin_content, withWeight=True, topK=2000)  # 得到每个单词的tfidf
workbook = Workbook()
sheet = workbook.active
for element in keywords:
    sheet.append(element)
workbook.save('data_tfidf.xlsx')  # 保存tfidf到xlsx
tfidf_dict = dict()
for element in keywords:
    tfidf_dict.update({element[0]: element[1]})  # 生成一个tfidf词典

# 存放每题中词按照tfidf值排序
sorted_words_list = list()
words_list = list()
for line in f.readlines():
    words = line.strip().split()
    words_list.extend(words)
    single_words = list(set(words))
    words_dict = dict()
    sorted_words = list()
    for word in single_words:
        words_dict.update({word: tfidf_dict.get(word, 0)})  # 建立每一句话中tfidf的值词典
    sorted_words = sorted(words_dict.items(), key=lambda x: x[1], reverse=True)  # 将每句话中的单词按照权重值进行排序
    sorted_words_list.append(sorted_words)
# 保存每题中词按照tfidf值排序到xlsx
workbook = Workbook()
sheet = workbook.active
for element in sorted_words_list:
    row = list()
    for key_value in element:
        row.append(key_value[0])
        row.append(key_value[1])
    sheet.append(row)
workbook.save('sorted_tfidf.xlsx')

# 单个词的词频
freq = Counter(words_list)
print('总词数为：', len(words_list), '\n单词数为：', len(freq))
key = list(freq.keys())
val = list(freq.values())
freq_list = list()
for i in range(len(freq)):
    element = list()
    element.append(key[i])
    element.append(val[i])
    freq_list.append(element)
# 将词频保存到excel表
workbook = Workbook()
sheet = workbook.active
for element in freq_list:
    sheet.append(element)
workbook.save('data_freq.xlsx')

f.close()


