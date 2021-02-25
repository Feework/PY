from openpyxl import load_workbook
from jieba import lcut
import re

# 结巴分词
# 读取提取出的词库 openpyxl模块是一个读写Excel 2010文档的Python库 load_workbook读取exel文件
workbook = load_workbook('data.xlsx')
sheet = workbook.get_sheet_by_name('Sheet')
rows = sheet.max_row
cols = sheet.max_column

# 加载停用词库
stop_f = open('stop_words.txt', "r", encoding='utf-8')
stop_words = list()
for line in stop_f.readlines():
    line = line.strip()
    if not len(line):
        continue
    stop_words.append(line)

# 使用 jieba 进行分词
str_list = list()
for row in range(1, rows+1):
    out_str = ''
    data_title = sheet.cell(row, 1).value
    data_ans = sheet.cell(row, 2).value
    word_list = lcut(data_title, cut_all=False) + lcut(data_ans, cut_all=False)
    #结巴分词 去掉停用词、百分数、小数、单字
    for word in word_list:
        if word not in stop_words:
            if re.search('([0-9.]+)%', word) == None and re.search('[0-9]{1,}[.][0-9]*', word) == None and len(word) != 1:
                if word != '\t':
                    out_str += word.lower()
                    out_str += ' '
    str_list.append(out_str)

# 将分词结果保存至txt文档，方便后面的提取和分析
with open("data_cut.txt", "w", encoding='utf-8') as fw:
    for element in str_list:
        element.encode('utf-8')
        data = element.strip()
        if len(data) != 0:
            fw.write(data)
            fw.write("\n")

stop_f.close()
fw.close()
